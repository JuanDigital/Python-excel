# -*- coding: utf-8 -*-
"""
cotiza
"""
import string
import openpyxl
import pandas as pd
import tkinter 
from tkinter import* 
from tkinter.ttk import*
import matplotlib.pyplot as plt
'''leer un libro xlsx'''

from openpyxl import load_workbook
book=load_workbook('bitacora.xlsx')
sheet=book.active
#wb = Workbook()
#ws = wb.active



titulos=['Biomedica S.A.C,COTIZACIÓN,6° Av. 91',
        'Calle El Amigo,  25-2-Ciudad: San Salvdor',
        'Sitio Web: tudominio.com,Teléfono: 2535-2575']

cotiza=[]
ncotiza=[]
ccotiza=['CLIENTE: ']
pcotiza=[]
bitacora=[]
#guia=''

def listAlphabet():
    return list(string.ascii_lowercase)
letras=(listAlphabet())
def estadisticaV():
    ventanaE=Tk()
    
    ventanaE.title('Elegir estadistica')
    ventanaE.geometry('800x800')
    #tot=Label(ventanaE,text = 'total')
    #tot.place(x=350,y=550)
    baseE=pd.ExcelFile('bitacora.xlsx')
    dfEsClient=pd.read_excel(baseE,'Sheet',index_col=False,usecols='A,B')
    dfMcli=pd.read_excel(baseE,'Sheet',index_col=False,usecols='A')
    dfprod=pd.read_excel(baseE,'Sheet',index_col=False,usecols='c:z')
    #print(dfprod)
    #dfprod.value_counts()
    #df3 = pd.DataFrame(dfprod)
    #df3.value_counts().plot(kind='bar')
    
    MejCli=dfEsClient['CLIENTE'].value_counts()#.plot(kind='bar')
    df2=pd.DataFrame(data=MejCli)
    #pic2=df2.plot.line()
    pic2=df2.plot(kind='bar',figsize=(10,25),grid=True)
    pic2=pic2.get_figure()
    pic2.savefig('cotizaciones por cliente.jpg')
    #MejCli.savefig("ClienteQueMasCotiza.jpg")
    #Resu=dfEsClient['CLIENTE'].value_counts()
    #print(Resu)
        
    #print(dfEsClient['CLIENTE'].value_counts())
    #dfEsClient['CLIENTE'].value_counts()
    #dfEsClient.plot.line(x='CLIENTE', y='TOTAL')
    
    df=pd.DataFrame(data=dfEsClient)
    pic=df.plot.line(title='Presupuesto de clientes',x='CLIENTE',figsize=(30,6));
    #pic=df.plot(kind='bar',x='CLIENTES')
    pic=pic.get_figure()
    pic.savefig('Presupuesto de clientes.jpg')
    
def ventaCotiza():
    ventana5=Tk()
    
    ventana5.title('cotización')
    ventana5.geometry('800x800')
    tot=Label(ventana5,text = 'total')
    tot.place(x=350,y=550)
    
    
    
    ''' crear cuadro de lista '''
    lista_elementos=Listbox(ventana5,width=50,height=3) 
    nombresB=Listbox(ventana5,width=100,height=2) 
    produc=Listbox(ventana5,width=50,height=20) 
    pres=Listbox(ventana5,width=50,height=20) 
    total=Listbox(ventana5,width=50,height=2) 
    ''' ubicacion de la lista'''
    lista_elementos.place(x=100,y=100)
    nombresB.place(x=100, y=150)
    produc.place(x=100, y=200)
    pres.place(x=400, y=200)
    total.place(x=400,y=550)
   
    
    '''insertando elementos '''
    def mostr():
        k=-1
        d=0        
        for a in titulos:
            #print (a,len(a))
            #imprime=(a)
            lista_elementos.insert(END,a)
        
        nombresB.insert(END,ccotiza)
        
        for b in ncotiza:
            produc.insert(END, b)
            
        for c in pcotiza:
            pres.insert(END, c)
            d+=int(c)
        bitacora.insert(1,c)    
        total.insert(END, d)
        
        '''donde escribir'''
            
        conta=open('i.txt','rt')
        i=conta.readline()
        #print(i)
        #print (type(i))# i=0, type= str
        j=int(i)+1
        #print (j)
       
      
        dor=open('i.txt','w')
        dor.write(str(j))
        dor.close()
        
        '''guardar datos en bitacora'''
    
       
       
        for ind in bitacora:
           k+=1
           guia=letras[k]+str(j)
           sheet[guia]=ind
        
        book.save('bitacora.xlsx')
        
        '''
     
        for abc in letras:
            k+=1
            guia=abc+(j)
            ws[guia]=str(bitacora[k])
       
        '''
        
        
        '''guardar libro'''
       # wb.save('bitacora.xlsx')
      
        
            
            
            
    botonv5= Button(ventana5, text='mostrar',command=mostr)
    botonv5.pack()
    #hola='hola'
    
        #lista_elementos.insert(END,b)
   
   # equipo=Entry(ventana4,font = 'Helvetica 20')
    #equipo.pack()
    
def ventaEquipo():
    ventana4=Tk()
    
    ventana4.title('Equipo')
    #ventana4.geometry('600x600')
    accion4=Label(ventana4,text = 'ingresa el numero correspondiente al equipo que desea cotizar')
    accion4.pack()
    
    equipo=Entry(ventana4,font = 'Helvetica 20')
    equipo.pack(side=BOTTOM)
    
    
    basex=pd.ExcelFile('base.xlsx',)
    dfequipo=pd.read_excel(basex,'Eq',usecols='b,c,i')
    
    bienvenida4=tkinter.Label(ventana4,text=dfequipo)
    bienvenida4.pack()
    
    def selEq():
        
        numequipo=int(equipo.get())+1
        
        selectEquipo=(dfequipo[int(equipo.get()):numequipo])
        
        lse=list(selectEquipo['EQUIPO'])
        nEq=str(lse[0])
        ncotiza.append(nEq)
        bitacora.append(nEq)
        
        lsep=list(selectEquipo['PRECIO'])
        nEqp=str(lsep[0])
        pcotiza.append(nEqp)
        
        #cotiza.append(selectEquipo)
        #print(cotiza)
       # ventaMobi()
        
    
    def SigVent4():
        ventaCotiza()
        ventana4.destroy()
      
    
    botonv4= Button(ventana4, text='agregar',command=selEq)
    botonv4.pack()
    botonv4_1= Button(ventana4, text='siguiente',command=SigVent4)
    botonv4_1.pack()
    
    
def ventaMobi():
    ventana3=Tk()
    ventana3.title('Mobiliario')
    #ventana3.geometry('600x600')
    accion3=Label(ventana3,text = 'ingresa el numero correspondiente al mobiliario que desea cotizar ')
    accion3.pack()
   # mob=StringVar()
    mob=Entry(ventana3,font = 'Helvetica 20')
    mob.pack(side=BOTTOM)
    
    
    basex=pd.ExcelFile('base.xlsx',)
    dfmobi=pd.read_excel(basex,'Mobiliario',usecols='c,d,e')
    #nombre=str(dfClient)
    bienvenida3=tkinter.Label(ventana3,text=dfmobi)
    bienvenida3.pack()
    
    def selMb():
        
        nummob=int(mob.get())+1
        
        selectMob=(dfmobi[int(mob.get()):nummob])
        
        lsm=list(selectMob['MOBILIARIO'])
        nMobi=str(lsm[0])
        ncotiza.append(nMobi)
        bitacora.append(nMobi)
        
        lsmp=list(selectMob['PRECIO'])
        nMobiP=str(lsmp[0])
        pcotiza.append(nMobiP)
        
        
       # print(cotiza)
       # ventaMobi()
        
    
    def SigVent3():
        ventaEquipo()
        ventana3.destroy()
        
      
    
    botonv3= Button(ventana3, text='agregar',command=selMb)
    botonv3.pack()
    botonv3_1= Button(ventana3, text='siguiente',command=SigVent3)
    botonv3_1.pack()
    
  
def NewVentana():
    #ventana.destroy()
    ventana2=tkinter.Tk()
    ventana2.title('CLIENTES')
    ventana2.geometry('600x600')
    accion=Label(ventana2,text = 'ingresa el numero correspondiente al cliente en el siguiente espacio')
    accion.pack()
    ncliente=StringVar()
    nclientes=Entry(ventana2,textvariable=ncliente,font = 'Helvetica 20')
    nclientes.pack(side=BOTTOM)
   
        
    
    '''
7    b=7
    c=b+1
'''
    basex=pd.ExcelFile('base.xlsx',)
    dfClient=pd.read_excel(basex,'Clientes',usecols='B,e,f')
    #nombre=str(dfClient)
    bienvenida2=tkinter.Label(ventana2,text=dfClient)
    bienvenida2.pack()
    
    def SigVent2():
        
        numcliente=int(nclientes.get())+1
        #print(dfClient[int(nclientes.get()):numcliente])
        selectCliente=(dfClient[int(nclientes.get()):numcliente])
        ''' obtener una cadena de un dataframe'''
        lsc=list(selectCliente['CLIENTE'])
        nCliente=str(lsc[0])
        ccotiza.append(nCliente)
        bitacora.insert(0, nCliente)
        #print(nCliente)  
        ccotiza.append('Tel.:')
        lscT=list(selectCliente['TELÉFONO'])
        ntCliente=str(lscT[0])
        
        #print(nCliente)
        ccotiza.append(ntCliente)
       # print(cotiza)
        ventaMobi()
        ventana2.destroy()
        
        '''lsc=list(selectCliente['CLIENTE'])
        nCliente=str(lsc[0])
        print(nCliente)'''
       # print(dfClient)]

       # print(ncliente.get())
        #numcliente=ncliente.get()
        #numcliente=int(numcliente)+1
        #print(numcliente)
    
    botonv2= Button(ventana2, text='ok',command=SigVent2)
    botonv2.pack()
    

def inicio():
    
    ventana=tkinter.Tk()
    ventana.title('Login')
    ventana.geometry('600x300')
    ventana.resizable(width=False, height=False)
    
    bienvenida=Label(ventana,text = 'Hola, ingresa para cotizar')
    bienvenida.pack()
    
    usuario=StringVar()
    user=Entry(ventana, textvariable=usuario, font = 'Helvetica 20')
    user.pack()
    
    contra=Label(ventana,text = 'contraseña')
    contra.pack()
    
    secret=StringVar()
    secreto=Entry(ventana, textvariable=secret, font = 'Helvetica 20',show='x')
    secreto.pack()
    
    def SigVent():
        if usuario.get()=='' and secret.get()=='':
            NewVentana()
            ventana.destroy()
            
        else:
            ventana.title('incorrecto')
        
    def Ventestad():
         if usuario.get()=='' and secret.get()=='':
             estadisticaV()
             
         else:
             ventana.title('ingresa con usuario y contraseña')
    
    
    boton1= Button(ventana, text='ACCESO',command=SigVent)
    boton1.pack(side= BOTTOM)
    estadistica=Button(ventana, text='Estadisticas',command=Ventestad).place(x=520,y=270)

    ventana.mainloop()
#                       ventaCotiza()
inicio()

  

    


'''
root =Tk()
root.title('login')
root.geometry('300X600')
root.resizable(width=False, heigth=False)
'''


#print(basex.sheet_names)
#dfus=basex.parse('Clientes')
#print(dfus)

#print(dfus.columns)

#colClient=[1,2,3,4,5]


